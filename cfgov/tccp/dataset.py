from django.utils.text import slugify

from openpyxl import load_workbook


def convert_header_to_field_name(s):
    return slugify(
        s.replace("$", "dollars").replace("%", "percentage")
    ).replace("-", "_")


def read_survey_data_from_worksheet(worksheet, start_row=9, start_column=1):
    headers = [
        cell.value
        for cell in worksheet[start_row + 1][start_column:]
        if cell.value
    ]

    return [
        {
            convert_header_to_field_name(header): row[i + start_column].value
            for i, header in enumerate(headers)
        }
        for row in worksheet.iter_rows(min_row=start_row + 2)
        if any(cell.value for cell in row)
    ]


def read_survey_data_from_stream(stream):
    workbook = load_workbook(stream, data_only=True, read_only=True)
    return read_survey_data_from_worksheet(workbook.worksheets[0])
